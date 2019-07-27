import openpyxl

#print(dir(openpyxl))

path = "C:\purushotham\Learning\python-learning\Python-practice\Excels\students.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

print((sheet_obj.cell(row=1,column=1)).value)

#-- Priniting maximum number of rows

print(sheet_obj.max_row)

#--- Printing maximum number of column

print(sheet_obj.max_column)


#-- Printing the column names

max_col = sheet_obj.max_column

for i in range(1, max_col+1):

    cell_obj = sheet_obj.cell(row=1,column=i)
    print(cell_obj.value)

#-- Printing the column value

max_row = sheet_obj.max_row

for i in range(1, max_row+1):
    cell_obj = sheet_obj.cell(row=i,column=1)
    print(cell_obj.value)

#-- printing the row values

for i in range(1,max_col+1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")

