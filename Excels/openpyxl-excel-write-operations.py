import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

print("sheet title is : ", sheet.title)

#-- Changing the sheet title

sheet.title = "openpyxl"

print("sheet title is : ", sheet.title)

#-- Writing to an excel sheet

#print((sheet.cell).__doc__)

c1 = sheet.cell(row=1,column=1)
c1.value = "Purushotham"

c2 = sheet.cell(row=1,column=2)
c2.value = "Reddy"

c3 = sheet['A2']
c3.value = "Chandra"

c4 = sheet['B2']
c4.value = "Reddy"

#--- creating new sheet in the workbook

wb.create_sheet(index=1,title="demo sheet")

wb.save("C:\purushotham\Learning\python-learning\Python-practice\Excels\openpyxl-read-write.xlsx")
