import openpyxl
from openpyxl.styles import Font
#-- setting the height & width of the cell

wb = openpyxl.Workbook()

sheet = wb.active

sheet.cell(row=1,column=1).value = "Hello"
sheet.row_dimensions[1].height = 70

sheet.cell(row=2,column=2).value = "World"
sheet.column_dimensions['B'].width = 40

wb.save("dimensions.xlsx")

#--  merging the cells ---

wb = openpyxl.Workbook()

sheet = wb.active

sheet.merge_cells('A2:D4')

sheet.cell(row=2,column=1).value="Hello World"

sheet.merge_cells('C6:D6')

sheet.cell(row=6,column=6).value="Welcome"

wb.save("merge.xlsx")

#------ unmerging the cells

wb = openpyxl.load_workbook("C:\purushotham\Learning\python-learning\Python-practice\Excels\merge.xlsx")

sheet = wb.active

sheet.unmerge_cells('A2:D4')
sheet.unmerge_cells('C6:D6')
wb.save("merge.xlsx")

#-- adding styles to the cell values

wb = openpyxl.Workbook()
sheet = wb.active

sheet.cell(row=1,column=1).value="Purushotham"
sheet.cell(row=1,column=1).font = Font(size=24)
sheet.cell(row=2,column=2).value="Chandra"
sheet.cell(row=2,column=2).font = Font(size=24,bold=True)
sheet.cell(row=3,column=3).value="Reddy"
sheet.cell(row=3,column=3).font = Font(size=24,italic=True)
sheet.cell(row=4,column=4).value="Hero"
sheet.cell(row=3,column=3).font = Font(size=24,name="Times New Roman")

wb.save("styles.xlsx")


